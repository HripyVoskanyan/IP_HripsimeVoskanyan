import os
import pandas as pd
from openpyxl import load_workbook

base_dir = 'C:/AUA/Image Processing/HW1/'
excel_path = os.path.join(base_dir, 'Hist_HSV.xlsx')

wb = load_workbook(excel_path)


folder_to_sheet = {
    'hue': 'Hist_Hue',
    'sat': 'Hist_Sat',
    'val': 'Hist_Val'
}

# Process each folder and corresponding sheet
for folder, sheet_name in folder_to_sheet.items():
    folder_path = os.path.join(base_dir, folder + '/')
    ws = wb[sheet_name]


    print(f"Values in cells from C2 to Z2 in {sheet_name}:")
    excel_headers = {}
    for col in range(3, 27):
        cell_value = ws.cell(row=2, column=col).value
        cell_coordinate = ws.cell(row=2, column=col).coordinate
        excel_headers[cell_value] = col  
        print(f"Cell {cell_coordinate}: {cell_value}")

    # Read filenames from the current directory
    csv_data = {}
    for filename in os.listdir(folder_path):
        if filename.endswith('.csv'):
            first_ten = filename[:10]
            csv_data[first_ten] = pd.read_csv(os.path.join(folder_path, filename), header=None)

    # Match and write data
    for header, col_index in excel_headers.items():
        first_ten_header = header[:10]
        if first_ten_header in csv_data:
            data = csv_data[first_ten_header].iloc[:, 0]  # Assuming data is in the first column
            for row_index, value in enumerate(data, start=3):  # Start writing from row 3
                ws.cell(row=row_index, column=col_index, value=value)

# Save and close the workbook
wb.save(excel_path)
wb.close()
