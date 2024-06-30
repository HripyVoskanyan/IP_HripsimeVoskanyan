import os
import pandas as pd
from openpyxl import load_workbook

def find_thresholds(histogram_values):
    values_list = histogram_values.tolist()
    peak_value = max(values_list)
    peak_index = values_list.index(peak_value)

    min_threshold = peak_index
    for i in range(peak_index, 0, -1):
        if values_list[i] < values_list[i - 1]:
            min_threshold = i
            break

    max_threshold = peak_index
    for i in range(peak_index, len(values_list) - 1):
        if values_list[i] < values_list[i + 1]:
            max_threshold = i
            break

    return min_threshold, max_threshold

# Define the base directory
base_dir = 'C:/AUA/Image Processing/HW1/'
# uncomment the other parts when doing Task 1, 2, 3
# uncomment the sv when doing Task 4
folder_to_sheet = {
    #'hue': 'Hist_Hue',
    #'sat': 'Hist_Sat',
    #'val': 'Hist_Val',
    'sv' : 'Hist_SV'
}

# Load the existing Excel workbook
wb = load_workbook(os.path.join(base_dir, 'Hist_HSV.xlsx'))

for folder, sheet_name in folder_to_sheet.items():
    folder_path = os.path.join(base_dir, folder + '/')
    ws = wb[sheet_name]

    print(f"Values in cells from C2 to Z2 in {sheet_name}:")
    excel_headers = {}
    for col in range(3, 27):  # Assuming headers are from C2 to Z2
        cell_value = ws.cell(row=2, column=col).value
        cell_coordinate = ws.cell(row=2, column=col).coordinate
        excel_headers[cell_value] = col
        print(f"Cell {cell_coordinate}: {cell_value}")

    # Process each histogram file in the directory
    for filename in os.listdir(folder_path):
        if filename.endswith('.csv'):
            # uncomment when doing Task 1, 2, 3

            #first_ten = filename[:10]
            # comment the previous line and uncomment this when doing Task 4
            first_ten = filename[3:9] + ".jpg"
            csv_file_path = os.path.join(folder_path, filename)
            histogram_data = pd.read_csv(csv_file_path, header=None).iloc[:, 0]

            min_thresh, max_thresh = find_thresholds(histogram_data)
            if first_ten in excel_headers:
                col_index = excel_headers[first_ten]
                ws.cell(row=260, column=col_index, value=min_thresh)
                ws.cell(row=261, column=col_index, value=max_thresh)
                print(f"Written min/max thresholds for {filename} to column {ws.cell(row=1, column=col_index).coordinate}")

# Save the workbook after all changes
wb.save(os.path.join(base_dir, 'Hist_HSV.xlsx'))
wb.close()
